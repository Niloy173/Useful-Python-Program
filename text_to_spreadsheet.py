# a single .txt file containg some data forwarding to spreadsheet in this program

import openpyxl,os

wb = openpyxl.Workbook()

sheet = wb.active

newsheet = wb.create_sheet(index = 0,title = "New")




os.chdir("E:/Useful programme") # .txt file contains this directory

row_ = 1
column_ =1

for file in os.listdir():

    if file.endswith('.txt'):

        with open(file) as f:

            for lines in f:

                newsheet.cell(row = row_,column = column_).value = lines
                row_ +=1

        column_ +=1

wb.save("text_to_spreadsheet.xlsx")

        
